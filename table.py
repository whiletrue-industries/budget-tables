import openpyxl
import decimal

BG_COLOR_NAMES = 'fabf8f'
BG_COLOR_HEADER = '95b3d7'


class Table():

    def __init__(self, title, group_fields=None, cleanup_fields=None):
        self.headers = dict()
        self.rows = list()
        self.row = dict()
        self.title = title
        self.groups = dict()
        self.group_fields = group_fields
        self.cleanup_fields = cleanup_fields
        self.alternating = 0

    def new_row(self, key):
        self.row = dict()
        self.rows.append((key, self.row))

    def set(self, key, value, header_order, **options):
        self.row[key] = dict(
            value=value,
            **options
        )
        if key not in self.headers:
            header_options = dict(**options)
            header_options.update(dict(
                bold=True,
                font_family='David Libre',
                font_size=12,
                color='FFFFFF',
                background_color='222446'
            ))
            self.headers[key] = dict(   
                value=key,
                score=header_order,
                **header_options
            )

    def get(self, key):
        return self.row.get(key, {}).get('value')
    
    def group(self, row, level, value):
        self.groups.setdefault(level, {}).setdefault(value, set()).add(row)

    def append_cells(self, recs):
        values = [x.get('value') if x else '' for x in recs]
        self.ws.append(values)
        row = self.ws[self.ws.max_row]
        reset_alternating = True
        for rec, cell in zip(recs, row):
            if rec is None:
                rec = dict(
                    background_color='ffffff'
                )
            if rec.get('number_format'):
                cell.number_format = rec['number_format']
            else:
                cell.number_format = "#,##0"
            align = rec.get('align', 'right') or 'right'
            cell.alignment = openpyxl.styles.Alignment(horizontal=align, vertical="center", wrapText=True, readingOrder=2)
            font_options = dict()
            if rec.get('bold'):
                font_options['bold'] = True
            if rec.get('color'):
                color = rec['color']
                if callable(color):
                    color = color(cell.value)
                if isinstance(color, list):
                    color = color[self.alternating]
                    reset_alternating = False
                font_options['color'] = color
            if rec.get('font_family'):
                font_options['name'] = rec.get('font_family')
            if rec.get('font_size'):
                font_options['size'] = rec.get('font_size')
            if font_options:
                cell.font = openpyxl.styles.Font(**font_options)
            if rec.get('border_bottom'):
                # add border below
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
            if rec.get('background_color'):
                if callable(rec['background_color']):
                    color = rec['background_color'](cell.value)
                else:
                    color = rec['background_color']
                if isinstance(color, list):
                    color = color[self.alternating]
                    reset_alternating = False
                cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
            elif not rec.get('parity'):
                cell.fill = openpyxl.styles.PatternFill(start_color="CAE9F5", end_color="CAE9F5", fill_type="solid")
            if rec.get('comment'):
                cell.comment = openpyxl.comments.Comment(rec['comment'], '-')
        return reset_alternating

    def process_value(self, value):
        if value is None:
            return ''
        if isinstance(value, float):
            return f'{value:,.2f}'
        if isinstance(value, int):
            return f'{value:}'
        return value


    def save(self, filename):
        headers = sorted(self.headers.values(), key=lambda x: x['score'])
        rows = sorted(self.rows, key=lambda x: x[0])
        print('TOTAL ROWS', len(rows))
        # row_to_ids = dict((x[0], i+2) for i, x in enumerate(rows))
        rows = [x[1] for x in rows]
        if self.cleanup_fields:
            running_header = [None] * len(self.cleanup_fields)
        for i, row in enumerate(rows):
            for v in row.values():
                if v.get('parity'):
                    v['parity'] = i % 2
            if self.group_fields:
                for l, f in enumerate(self.group_fields):
                    if f in row and row[f]['value']:
                        self.group(i+2, l+1, row[f]['value'])
            if self.cleanup_fields:
                for j, f in enumerate(self.cleanup_fields):
                    if f in row and row[f]['value']:
                        if running_header[j] != row[f]['value']:
                            running_header[j] = row[f]['value']
                            for k in range(j+1, len(self.cleanup_fields)):
                                running_header[k] = None
                        else:
                            row[f]['value'] = ''

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.append_cells(headers)
        self.ws.title = self.title
        self.ws.sheet_view.rightToLeft = True
        for row in rows:
            processed_row = []
            for h in headers:
                rec = row.get(h['value'])
                if rec is not None:
                    rec['value'] = self.process_value(rec['value'])
                processed_row.append(rec)
            if self.append_cells(processed_row):
                self.alternating = 0
            else:
                self.alternating = 1 - self.alternating
            # print(row)
            # assert False
                
        # Make sure all columns are wide enough
        for column in self.ws.columns:
            column = list(column)
            header = None
            if len(column) > 1:
                header = column[0].value
                column = column[1:]
            max_length = max(
                (len(f"{cell.value:,.1f}") if isinstance(cell.value, decimal.Decimal) else len(str(cell.value)))
                for cell in column)
            if header:
                header = header.split()
                max_length = max(max_length, max(len(x) for x in header))
            adjusted_width = (max_length + 2) * 1.0
            adjusted_width = min(adjusted_width, 175/7)
            self.ws.column_dimensions[column[0].column_letter].width = adjusted_width

        for level in sorted(self.groups.keys()):
            print('GROUPING', level, self.groups.keys())
            self.ws.sheet_properties.outlinePr.summaryBelow = False
            for k, group in self.groups[level].items():
                # print('GROUPING: ', level, k, group)
                assert len(group) == (max(group) - min(group) + 1)
                if len(group) > 1:
                    self.ws.row_dimensions.group(min(group)+1, max(group), outline_level=level, hidden=level==3)

        self.ws.freeze_panes = self.ws['A2']

        self.wb.save(filename)

def color_scheme_red_green(default_color='FFFFFF'):
    def func(value):
        if not value:
            return default_color
        value = float(value)
        if value > .05:
            return 'FF6432'
        if value < -.05:
            return 'cbf99f'
        return default_color
    return func