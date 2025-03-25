import openpyxl
import dataflows as DF
import decimal

BUDGET_SOURCE_DATAPACKAGE = 'https://next.obudget.org/datapackages/budget/national/processed/with-extras/datapackage.json'
CONNECTED_SOURCE_DATAPACKAGE = 'https://next.obudget.org/datapackages/budget/national/processed/connected-items-explained/datapackage.json'
MIN_YEAR = 2021

CHECKPOINT_DIR = '.checkpoints/proposal-compare'

BG_COLOR_NAMES = 'fabf8f'
BG_COLOR_HEADER = '95b3d7'


def nice_code(code):
    return code[2:]
    # With dots:
    # ret = []
    # code = code[2:]
    # return f'="{code}"'
    # while code:
    #     ret.append(code[:2])
    #     code = code[2:]
    # return '.'.join(ret)

def check_for_active(row):
    if (row['net_allocated'] or row['net_revised'] or row['net_executed']):
        return True
    history = row.get('history') or {}
    for year, rec in history.items():
        year = int(year)
        if year >= MIN_YEAR:
            if rec.get('net_allocated') or rec.get('net_revised') or rec.get('net_executed'):
                return True
    return False

def get_proposal_data():
    raw = DF.Flow(
        DF.load(BUDGET_SOURCE_DATAPACKAGE),
        DF.filter_rows(lambda row: row['year'] >= MIN_YEAR),
        DF.filter_rows(lambda row: not row['code'].startswith('0000')),
        DF.filter_rows(lambda row: not row['code'].startswith('C')),
        DF.filter_rows(lambda row: row['code'] < '0089'),
        DF.filter_rows(lambda row: len(row['code']) >= 4),
        DF.select_fields([
            'code',
            'year',
            'title',
            'non_repeating',
            'net_allocated',
            'net_revised',
            'net_executed',
        ]),
        DF.set_type('code', transform=nice_code),
        DF.checkpoint('raw', CHECKPOINT_DIR),
    ).results()[0][0]
    raw_map = dict(
        ((row['year'], row['code']), row)
        for row in raw
    )
    connected = DF.Flow(
        DF.load(CONNECTED_SOURCE_DATAPACKAGE),
        DF.filter_rows(lambda row: row['year'] >= MIN_YEAR),
        DF.filter_rows(lambda row: not row['code'].startswith('0000')),
        DF.filter_rows(lambda row: not row['code'].startswith('C')),
        DF.filter_rows(lambda row: row['code'] < '0089'),
        DF.filter_rows(lambda row: len(row['code']) >= 4),
        DF.filter_rows(check_for_active),
        # DF.filter_rows(lambda row: (len(row['code']) < 10) or (row['net_allocated'] or row['net_revised'] or row['net_executed'])),
        DF.select_fields([
            'code',
            'year',
            'title',
            'history',
            'hierarchy',
            'is_proposal',
        ]),
        DF.set_type('code', transform=nice_code),
        DF.checkpoint('connected', CHECKPOINT_DIR),
    ).results()[0][0]
    return raw_map, connected

class Table():

    def __init__(self, title, group_fields=None, cleanup_fields=None):
        self.headers = dict()
        self.rows = list()
        self.row = dict()
        self.title = title
        self.groups = dict()
        self.group_fields = group_fields
        self.cleanup_fields = cleanup_fields

    def new_row(self, key):
        self.row = dict()
        self.rows.append((key, self.row))

    def set(self, key, value, header_order, **options):
        self.row[key] = dict(
            value=value,
            **options
        )
        if key not in self.headers:
            header_options = dict(**options)
            header_options['bold'] = True
            header_options['border_bottom'] = True
            header_options['background_color'] = BG_COLOR_HEADER
            self.headers[key] = dict(
                value=key,
                score=header_order,
                **header_options
            )

    def get(self, key):
        return self.row.get(key, {}).get('value')
    
    def group(self, row, level, value):
        self.groups.setdefault(level, {}).setdefault(value, set()).add(row)

    def append_cells(self, recs):
        values = [x.get('value') if x else '' for x in recs]
        self.ws.append(values)
        row = self.ws[self.ws.max_row]
        for rec, cell in zip(recs, row):
            if rec is None:
                rec = dict(
                    background_color='cccccc'
                )
            if rec.get('number_format'):
                cell.number_format = rec['number_format']
            else:
                cell.number_format = "#,##0"
            cell.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center", wrap_text=False, readingOrder=2)
            if rec.get('bold'):
                cell.font = openpyxl.styles.Font(bold=True)
            if rec.get('border_bottom'):
                # add border below
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
            if rec.get('background_color'):
                if callable(rec['background_color']):
                    color = rec['background_color'](cell.value)
                else:
                    color = rec['background_color']
                cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
            elif not rec.get('parity'):
                cell.fill = openpyxl.styles.PatternFill(start_color="CAE9F5", end_color="CAE9F5", fill_type="solid")
            if rec.get('comment'):
                cell.comment = openpyxl.comments.Comment(rec['comment'], '-')

    def save(self, filename):
        headers = sorted(self.headers.values(), key=lambda x: x['score'])
        rows = sorted(self.rows, key=lambda x: x[0])
        print('TOTAL ROWS', len(rows))
        # row_to_ids = dict((x[0], i+2) for i, x in enumerate(rows))
        rows = [x[1] for x in rows]
        if self.cleanup_fields:
            running_header = [None] * len(self.cleanup_fields)
        for i, row in enumerate(rows):
            for v in row.values():
                if v.get('parity'):
                    v['parity'] = i % 2
            if self.group_fields:
                for l, f in enumerate(self.group_fields):
                    if f in row and row[f]['value']:
                        self.group(i+2, l+1, row[f]['value'])
            if self.cleanup_fields:
                for j, f in enumerate(self.cleanup_fields):
                    if f in row and row[f]['value']:
                        if running_header[j] != row[f]['value']:
                            running_header[j] = row[f]['value']
                            for k in range(j+1, len(self.cleanup_fields)):
                                running_header[k] = None
                        else:
                            row[f]['value'] = ''

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.append_cells(headers)
        self.ws.title = self.title
        self.ws.sheet_view.rightToLeft = True
        for row in rows:
            processed_row = []
            for h in headers:
                rec = row.get(h['value'])
                if rec is not None:
                    rec['value'] = process_value(rec['value'])
                processed_row.append(rec)
            self.append_cells(processed_row)
            # print(row)
            # assert False
                
        # Make sure all columns are wide enough
        for column in self.ws.columns:
            max_length = max(
                (len(f"{cell.value:,.1f}") if isinstance(cell.value, decimal.Decimal) else len(str(cell.value)))
                for cell in column)
            adjusted_width = (max_length + 2) * 1.0
            adjusted_width = min(adjusted_width, 175/7)
            self.ws.column_dimensions[column[0].column_letter].width = adjusted_width

        for level in sorted(self.groups.keys()):
            print('GROUPING', level, self.groups.keys())
            self.ws.sheet_properties.outlinePr.summaryBelow = False
            for k, group in self.groups[level].items():
                # print('GROUPING: ', level, k, group)
                assert len(group) == (max(group) - min(group) + 1)
                if len(group) > 1:
                    self.ws.row_dimensions.group(min(group)+1, max(group), outline_level=level, hidden=level==3)

        self.ws.freeze_panes = self.ws['A2']

        self.wb.save(filename)


def process_data():
    raw_map, connected = get_proposal_data()
    max_year = max(r['year'] for r in connected)
    proposal_year = list(set(r['year'] for r in connected if r['is_proposal']))
    proposal_year = proposal_year[0] if len(proposal_year) == 1 else None

    if proposal_year is None:
        raise Exception('Could not find a single proposal year, bailing out')
    before_proposal_year = proposal_year - 1

    used_keys = set()
    # table_rows = list()
    # headers = dict()

    titles_for_code_aux = dict()
    for item in raw_map.values():
        if item.get('net_allocated') or item.get('net_revised') or item.get('net_executed'):
            titles_for_code_aux.setdefault(item['code'], dict()).setdefault(item['title'], set()).add(item['year'])
    for k in titles_for_code_aux:
        titles_for_code_aux[k] = sorted((max(y), t) for t, y in titles_for_code_aux[k].items())
    titles_for_code = dict(
        (k, v[-1])
        for k, v in titles_for_code_aux.items()
    )
    titles_comments_for_code = dict()
    for k, v in titles_for_code_aux.items():
        if len(v) > 1:
            titles_comments_for_code[k] = ', '.join(f'עד שנת {y} נקרא {t}' for y, t in v[:-1])
            print('TFC', k, v)

    table = Table('השוואת הצעת התקציב', 
                  group_fields=['קוד סעיף', 'קוד תחום', 'קוד תכנית'],
                  cleanup_fields=['קוד סעיף', 'שם סעיף', 'קוד תחום', 'שם תחום', 'קוד תכנית', 'שם תכנית'])

    for year in range(max_year, MIN_YEAR-1, -1):
        print(f'PROCESSING YEAR {year}, got so far {len(used_keys)} keys')
        for item in connected:
            if item['year'] != year:
                continue
            code = item['code']
            title = item['title']
            # max_year_, title = titles_for_code[code]
            # if item['year'] != max_year:
            #     print(f'Item {code} is not max year {item["year"]} != {max_year}')
            #     continue
            # print(f'Processing {year} {code}')
            key = (year, code)
            if key in used_keys:
                continue
            keys = [(year, [code])]
            history = item['history']
            if history is not None:
                history = list((int(k), v) for k, v in history.items())
                history = sorted(history, key=lambda x: x[0], reverse=True)
                for _year, _rec in history:
                    if _year < MIN_YEAR:
                        break
                    if _rec.get('net_allocated') or _rec.get('net_revised') or _rec.get('net_executed'):
                        _year_codes = [nice_code(x.split(':')[0]) for x in _rec['code_titles']]
                        keys.append((_year, _year_codes))
            # print('KEYS', keys)

            hierarchy = item['hierarchy']
            code_titles = [(nice_code(h[0]), titles_for_code[nice_code(h[0])][1], None) for h in hierarchy[1:]] + [(code, title, titles_comments_for_code.get(code))]

            row_key = (code, year)
            table.new_row(row_key)

            table.set('קוד סעיף', '', 0, background_color=BG_COLOR_NAMES)
            table.set('שם סעיף', '', 1, background_color=BG_COLOR_NAMES)
            table.set('קוד תחום', '', 10, background_color=BG_COLOR_NAMES)
            table.set('שם תחום', '', 11, background_color=BG_COLOR_NAMES)
            table.set('קוד תכנית', '', 20, background_color=BG_COLOR_NAMES)
            table.set('שם תכנית', '', 21, background_color=BG_COLOR_NAMES)
            table.set('קוד תקנה', '', 30, background_color=BG_COLOR_NAMES)
            table.set('שם תקנה', '', 31, background_color=BG_COLOR_NAMES)

            # print('CCCC', code_titles)
            if len(code_titles) > 0:
                _code, _title, _comment = code_titles.pop(0)
                table.set('קוד סעיף', f'="{_code}"', 0, background_color=BG_COLOR_NAMES)
                table.set('שם סעיף', _title, 1, background_color=BG_COLOR_NAMES, comment=_comment)
            if len(code_titles) > 0:
                _code, _title, _comment = code_titles.pop(0)
                table.set('קוד תחום', f'="{_code}"', 10, background_color=BG_COLOR_NAMES)
                table.set('שם תחום', _title, 11, background_color=BG_COLOR_NAMES, comment=_comment)
            if len(code_titles) > 0:
                _code, _title, _comment = code_titles.pop(0)
                table.set('קוד תכנית', f'="{_code}"', 20, bold=True, background_color=BG_COLOR_NAMES)
                table.set('שם תכנית', _title, 21, bold=True, background_color=BG_COLOR_NAMES, comment=_comment)
            if len(code_titles) > 0:
                _code, _title, _comment = code_titles.pop(0)
                table.set('קוד תקנה', f'="{_code}"', 30, background_color=BG_COLOR_NAMES)
                table.set('שם תקנה', _title, 31, background_color=BG_COLOR_NAMES, comment=_comment)
            for _year, _codes in keys:
                sum_allocated = None
                sum_revised = None
                sum_executed = None
                titles = []
                codes = []
                for _code in _codes:
                    if (_year, _code) in used_keys:
                        continue

                    # print(f'Processing inner {_year} {_code}')
                    if (_year, _code) in raw_map:
                        raw = raw_map[(_year, _code)]
                        titles.append(raw['title'])
                        codes.append(_code)
                        if raw['net_allocated'] is not None:
                            sum_allocated = (sum_allocated or 0) + raw['net_allocated']
                        if raw['net_revised'] is not None:
                            sum_revised = (sum_revised or 0) + raw['net_revised']
                        if raw['net_executed'] is not None:
                            sum_executed = (sum_executed or 0) + raw['net_executed']
                        used_keys.add((_year, _code))
                    else:
                        print(f'Could not find {_year} {_code}')

                    options = dict(
                        bold=_year == max_year,
                        parity=True,
                        number_format='#,##0.0'
                    )
                codes = ', '.join(codes)
                titles = ', '.join(titles)
                if codes != code or titles != title:
                    options['comment'] = f'בשנת {_year} הסעיף נקרא {titles} - {codes}'
                if sum_allocated is not None:
                    if sum_allocated == 0 and _year != max_year:
                        comment = options.get('comment') or ''
                        if comment:
                            comment += '\n'
                        if sum_revised:
                            comment += f':התקציב המאושר {sum_revised:,.0f}'
                        elif sum_executed:
                            comment += f':התקציב המבוצע {sum_executed:,.0f}'
                        options_ = dict(options, comment=comment)
                    else:
                        options_ = options
                    table.set(f'{_year}', sum_allocated/1000000, _year*100 + 1, **options_)
                if _year == before_proposal_year:
                    if sum_revised is not None:
                        table.set(f'{_year} מאושר', sum_revised/1000000, _year*100 + 2, **options)
                    if sum_executed is not None:
                        table.set(f'{_year} מבוצע', sum_executed/1000000, _year*100 + 3, **options)
            
            max_year_allocated = table.get(f'{max_year}')
            before_proposal_year_allocated = table.get(f'{before_proposal_year}')
            before_proposal_year_revised = table.get(f'{before_proposal_year} מאושר')
            if None not in (max_year_allocated, before_proposal_year_allocated) and before_proposal_year_allocated > 0:
                change = (max_year_allocated - before_proposal_year_allocated) / before_proposal_year_allocated
                change = round(change, 2)
                table.set('שינוי מול מקורי 2024', change, (max_year+1)*100 + 1,
                    bold=False,
                    parity=1,
                    background_color=color_scheme_red_green,
                    number_format='0%'
                )
            if None not in (max_year_allocated, before_proposal_year_revised) and before_proposal_year_revised > 0:
                change = (max_year_allocated - before_proposal_year_revised) / before_proposal_year_revised
                change = round(change, 2)
                table.set('שינוי מול מאושר 2024', change, (max_year+1)*100 + 2,
                    bold=False,
                    parity=1,
                    background_color=color_scheme_red_green,
                    number_format='0%'
                )
    
    table.save('proposal-compare.xlsx')

def process_value(value):
    if value is None:
        return ''
    if isinstance(value, float):
        return f'{value:,.2f}'
    if isinstance(value, int):
        return f'{value:}'
    return value

def color_scheme_red_green(value):
    if not value:
        return 'FFFFFF'
    value = float(value)
    # if value < 0:
    #     value = min(-value, 1.0)
    #     value = 0xFF - int(value * (0xFF - 0x88)) 
    #     # print('FF{:02x}88'.format(value))
    #     return f'FF{value:02X}{value:02X}'
    # if value > 0:
    #     value = min(value, 1.0)
    #     value = 0xFF - int(value * (0xFF - 0x88))
    #     # print('88FF{:02x}'.format(value))
    #     return f'{value:02X}FF{value:02X}'.format(value)
    if value > .05:
        return 'FF0000'
    if value < -.05:
        return '00FF00'
    return 'FFFFFF'


if __name__=='__main__':
    process_data()
